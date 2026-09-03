from datetime import date, time, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.academics.models import Course, CourseClass
from apps.academics.services import enroll_student
from apps.billing.services import collect_payment, refund_payment, revenue_in_year
from apps.core.models import PaymentMethod
from apps.students.models import Student


class ReportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="admin",
            password="secure-test-password",
        )
        self.client.force_login(self.user)
        self.course = Course.objects.create(
            name="English Level 1",
            fee_type=Course.FeeType.MONTHLY,
            default_fee=Decimal("30.00"),
            currency="USD",
        )
        self.course_class = CourseClass.objects.create(
            course=self.course,
            name="English Level 1 – Morning",
            start_date=date(2026, 8, 1),
            study_days=[0, 2, 4],
            start_time=time(8, 0),
            end_time=time(9, 30),
        )
        self.student = Student.objects.create(
            name_kh="សុខា",
            name_en="Sokha",
            gender=Student.Gender.MALE,
            phone="012345678",
        )
        self.enrollment = enroll_student(
            self.student,
            self.course_class,
            user=self.user,
            next_due_date=timezone.localdate() - timedelta(days=2),
        )
        self.cash = PaymentMethod.objects.get(code="cash")

    def _pay(self, **overrides):
        today = timezone.localdate()
        payload = {
            "enrollment": self.enrollment,
            "paid_on": today,
            "tuition_amount": Decimal("30.00"),
            "discount_amount": Decimal("0.00"),
            "method": self.cash,
            "next_due_date": today + timedelta(days=30),
            "user": self.user,
        }
        payload.update(overrides)
        return collect_payment(**payload)

    def test_hub_and_sidebar(self):
        response = self.client.get(reverse("reports:index"))
        self.assertContains(response, "របាយការណ៍")
        self.assertContains(response, "ចំណូលថ្ងៃនេះ")
        self.assertContains(response, "សិស្សហួស Due Date")
        nav = self.client.get(reverse("dashboard:index"))
        self.assertContains(nav, reverse("reports:index"))
        self.assertContains(nav, "របាយការណ៍")
        self.assertContains(nav, 'href="/reports/"')

    def test_revenue_report_lists_payment_and_serial(self):
        payment = self._pay()
        response = self.client.get(reverse("reports:detail", args=["revenue"]))
        self.assertContains(response, "ល.រ")
        self.assertContains(response, "សុខា")
        self.assertContains(response, payment.total_display)
        self.assertContains(response, "Excel")
        self.assertContains(response, "PDF")
        self.assertContains(response, "date-field")
        self.assertContains(response, "ពីថ្ងៃ")
        self.assertContains(response, "ដល់ថ្ងៃ")
        html = response.content.decode()
        from_pos = html.find("id_date_from")
        to_pos = html.find("id_date_to")
        course_pos = html.find("id_course")
        self.assertLess(from_pos, to_pos)
        self.assertLess(to_pos, course_pos)
        self.assertIn("flex flex-wrap items-end", html)
        self.assertIn("w-52 max-w-full", html)
        self.assertNotIn("xl:grid-cols-4", html)

    def test_excel_export_contains_amount(self):
        self._pay()
        response = self.client.get(reverse("reports:excel", args=["revenue"]))
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]
        self.assertTrue(any(value and "30" in str(value) for value in values))
        self.assertTrue(any(value == "ល.រ" for value in values))

    def test_unpaid_and_overdue_reports(self):
        unpaid = self.client.get(reverse("reports:detail", args=["unpaid"]))
        overdue = self.client.get(reverse("reports:detail", args=["overdue"]))
        self.assertContains(unpaid, "សុខា")
        self.assertContains(overdue, "សុខា")
        self.assertContains(unpaid, "ល.រ")

    def test_paid_and_refund_reports(self):
        payment = self._pay()
        paid = self.client.get(reverse("reports:detail", args=["paid"]))
        self.assertContains(paid, "សុខា")
        refund_payment(payment, method=self.cash, reason="ឈប់រៀន", user=self.user)
        refunds = self.client.get(reverse("reports:detail", args=["refunds"]))
        self.assertContains(refunds, "ឈប់រៀន")
        self.assertContains(refunds, "$30.00")

    def test_revenue_nets_refunds_and_keeps_currencies_separate(self):
        payment = self._pay()
        today = timezone.localdate()
        refund_payment(payment, method=self.cash, reason="ឈប់រៀន", refunded_on=today, user=self.user)
        self.assertEqual(revenue_in_year(today, "USD"), Decimal("0.00"))
        response = self.client.get(reverse("reports:detail", args=["revenue"]))
        self.assertContains(response, "$0.00")
        self.assertContains(response, "៛0")

    def test_filter_by_class(self):
        self._pay()
        other = CourseClass.objects.create(
            course=self.course,
            name="English Level 1 – Evening",
            start_date=date(2026, 8, 1),
            study_days=[1, 3],
            start_time=time(17, 0),
            end_time=time(18, 30),
        )
        response = self.client.get(
            reverse("reports:detail", args=["revenue"]),
            {"course_class": other.pk},
        )
        self.assertContains(response, "មិនមានទិន្នន័យត្រូវនឹងតម្រង។")

    def test_unknown_kind_is_404(self):
        response = self.client.get(reverse("reports:detail", args=["not-a-kind"]))
        self.assertEqual(response.status_code, 404)

    def test_every_report_kind_loads(self):
        for kind in ("revenue", "paid", "unpaid", "overdue", "refunds", "attendance"):
            with self.subTest(kind=kind):
                response = self.client.get(reverse("reports:detail", args=[kind]))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "ល.រ")

    def test_attendance_report_lists_marked_student(self):
        from apps.academics.attendance import mark_class_attendance
        from apps.academics.models import AttendanceRecord

        mark_class_attendance(
            self.course_class,
            timezone.localdate(),
            {self.enrollment.pk: AttendanceRecord.Status.PRESENT},
            user=self.user,
        )
        response = self.client.get(reverse("reports:detail", args=["attendance"]))
        self.assertContains(response, "សុខា")
        self.assertContains(response, "វត្តមាន")
        self.assertContains(response, "យឺត")
        self.assertContains(response, "English Level 1 – Morning")

    def test_pdf_or_print_fallback(self):
        self._pay()
        response = self.client.get(reverse("reports:pdf", args=["revenue"]))
        if response.status_code == 200:
            self.assertEqual(response["Content-Type"], "application/pdf")
        else:
            self.assertEqual(response.status_code, 302)
