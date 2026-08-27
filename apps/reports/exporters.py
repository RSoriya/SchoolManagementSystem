from io import BytesIO

from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def excel_response(filename, sheets):
    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1557AB")
    title_font = Font(bold=True, size=14)

    first = True
    for sheet in sheets:
        if first:
            ws = workbook.active
            ws.title = sheet["title"][:31]
            first = False
        else:
            ws = workbook.create_sheet(sheet["title"][:31])
        row_num = 1
        if sheet.get("heading"):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(sheet["headers"]), 1))
            cell = ws.cell(1, 1, sheet["heading"])
            cell.font = title_font
            row_num = 2
            if sheet.get("subtitle"):
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(sheet["headers"]), 1))
                ws.cell(2, 1, sheet["subtitle"])
                row_num = 3
        for col, header in enumerate(sheet["headers"], start=1):
            cell = ws.cell(row_num, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in sheet["rows"]:
            row_num += 1
            for col, value in enumerate(row, start=1):
                ws.cell(row_num, col, value)
        for col in range(1, len(sheet["headers"]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        ws.auto_filter.ref = f"A{3 if sheet.get('heading') else 1}:{get_column_letter(len(sheet['headers']))}{row_num}"
        ws.freeze_panes = f"A{4 if sheet.get('heading') else 2}"

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def pdf_response(request, template_name, context, filename):
    html = render_to_string(template_name, context, request=request)
    try:
        from weasyprint import HTML

        pdf = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except Exception:
        return None
